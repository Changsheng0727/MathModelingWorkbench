from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from app.services.executor import run_python_script
from app.services.store import save_json


def generate_specialized_script(root: Path, analysis: dict[str, Any]) -> dict[str, str]:
    code_dir = root / "code"
    results_dir = root / "results"
    code_dir.mkdir(parents=True, exist_ok=True)
    results_dir.mkdir(parents=True, exist_ok=True)

    recommended = analysis.get("recommended_problem", {})
    payload = {
        "problem_id": recommended.get("id", "Unknown"),
        "problem_title": recommended.get("title", "Unknown problem"),
        "model_types": recommended.get("model_types", []),
        "tasks": recommended.get("tasks", []),
        "suggested_methods": recommended.get("suggested_methods", []),
    }
    script_path = code_dir / "run_specialized_model.py"
    script_path.write_text(render_specialized_script(payload), encoding="utf-8")
    return {"specialized_script": "code/run_specialized_model.py"}


def run_specialized_script(root: Path, timeout: int = 360) -> dict[str, Any]:
    script_path = root / "code" / "run_specialized_model.py"
    if not script_path.exists():
        raise FileNotFoundError("请先生成专项建模脚本")

    run_result = run_python_script(root, "code/run_specialized_model.py", "artifacts/specialized_run.log", timeout)
    manifest_path = root / "results" / "specialized_manifest.json"
    manifest: dict[str, Any] = {}
    if manifest_path.exists():
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        except Exception:
            manifest = {}
    payload = {
        "success": run_result["success"],
        "returncode": run_result["returncode"],
        "executor": run_result["executor"],
        "log": run_result["log"],
        "manifest": "results/specialized_manifest.json" if manifest_path.exists() else "",
        "outputs": manifest,
    }
    save_json(root / "artifacts" / "specialized_status.json", payload)
    return payload


def render_specialized_script(payload: dict[str, Any]) -> str:
    payload_json = json.dumps(payload, ensure_ascii=False, indent=2)
    template = r'''from __future__ import annotations

import itertools
import json
import math
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.ensemble import HistGradientBoostingRegressor, RandomForestRegressor
from sklearn.linear_model import Ridge
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.pipeline import make_pipeline
from sklearn.preprocessing import StandardScaler


CONFIG = __CONFIG_JSON__
ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "raw"
RESULTS_DIR = ROOT / "results"
TABLE_DIR = RESULTS_DIR / "tables"
FIG_DIR = RESULTS_DIR / "figures"
SPECIAL_DIR = RESULTS_DIR / "specialized"
PROBLEM_ID = CONFIG.get("problem_id", "Unknown")
PROBLEM_TITLE = CONFIG.get("problem_title", "Unknown problem")

plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "Arial Unicode MS", "DejaVu Sans"]
plt.rcParams["axes.unicode_minus"] = False
plt.rcParams["figure.facecolor"] = "white"
plt.rcParams["axes.facecolor"] = "white"
plt.rcParams["savefig.facecolor"] = "white"


def ensure_dirs() -> None:
    for path in [RESULTS_DIR, TABLE_DIR, FIG_DIR, SPECIAL_DIR]:
        path.mkdir(parents=True, exist_ok=True)


def read_csv(path: Path) -> pd.DataFrame:
    for encoding in ["utf-8-sig", "gbk", "utf-8"]:
        try:
            return pd.read_csv(path, encoding=encoding)
        except Exception:
            continue
    raise ValueError(f"cannot read csv: {path}")


def candidate_files() -> list[Path]:
    all_files = [p for p in sorted(RAW_DIR.rglob("*")) if p.is_file() and p.suffix.lower() in [".csv", ".xlsx", ".xls"]]
    selected = []
    for path in all_files:
        rel = path.relative_to(RAW_DIR).as_posix()
        if PROBLEM_ID != "Unknown" and (
            f"赛题{PROBLEM_ID}" in rel
            or f"Problem {PROBLEM_ID}" in rel
            or f"题{PROBLEM_ID}" in rel
            or f"{PROBLEM_ID}-" in rel
            or f"{PROBLEM_ID}_" in rel
        ):
            selected.append(path)
    return selected or all_files


def load_tables() -> list[tuple[str, pd.DataFrame]]:
    tables = []
    for path in candidate_files():
        rel = path.relative_to(RAW_DIR).as_posix()
        try:
            if path.suffix.lower() == ".csv":
                tables.append((rel, read_csv(path)))
            else:
                wb = load_workbook(path, read_only=True, data_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    if ws.max_row <= 1:
                        continue
                    nrows = min(ws.max_row, 200000)
                    df = pd.read_excel(path, sheet_name=sheet, nrows=nrows)
                    tables.append((f"{rel}::{sheet}", df))
                wb.close()
        except Exception as exc:
            print(f"[WARN] cannot load {rel}: {type(exc).__name__}: {exc}")
    return tables


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    frame = df.copy()
    frame.columns = [str(c).strip() for c in frame.columns]
    return frame


def find_datetime_col(df: pd.DataFrame) -> str | None:
    cols = list(df.columns)
    preferred = []
    for col in cols:
        low = str(col).lower()
        if any(key in low for key in ["time", "date", "日期", "时间"]):
            preferred.append(col)
    preferred.extend([c for c in cols if c not in preferred])
    sample = df.head(3000)
    for col in preferred:
        parsed = pd.to_datetime(sample[col], errors="coerce")
        if parsed.notna().mean() >= 0.6:
            return str(col)
    return None


def merge_compatible_tables(tables: list[tuple[str, pd.DataFrame]]) -> list[tuple[str, pd.DataFrame]]:
    groups: dict[tuple[str, ...], list[pd.DataFrame]] = {}
    names: dict[tuple[str, ...], list[str]] = {}
    for name, df in tables:
        df = normalize_columns(df)
        key = tuple(df.columns)
        groups.setdefault(key, []).append(df)
        names.setdefault(key, []).append(name)
    merged = []
    for key, frames in groups.items():
        title = " + ".join(names[key][:3])
        if len(names[key]) > 3:
            title += f" 等{len(names[key])}表"
        merged.append((title, pd.concat(frames, ignore_index=True)))
    return merged


def build_daily_dataset(tables: list[tuple[str, pd.DataFrame]]) -> tuple[str, pd.DataFrame] | None:
    candidates = []
    for name, df in merge_compatible_tables(tables):
        dt_col = find_datetime_col(df)
        if not dt_col:
            continue
        numeric_cols = []
        for col in df.columns:
            if str(col) == str(dt_col):
                continue
            converted = pd.to_numeric(df[col], errors="coerce")
            if converted.notna().mean() >= 0.55:
                numeric_cols.append(str(col))
        if not numeric_cols:
            continue
        score = len(df) * (1 + min(len(numeric_cols), 8))
        candidates.append((score, name, dt_col, numeric_cols, df))
    if not candidates:
        return None
    _, name, dt_col, numeric_cols, df = max(candidates, key=lambda x: x[0])
    work = df.copy()
    work[dt_col] = pd.to_datetime(work[dt_col], errors="coerce")
    work = work[work[dt_col].notna()]
    work["date"] = work[dt_col].dt.normalize()
    daily = work.groupby("date").size().rename("record_count").to_frame()
    for col in numeric_cols[:15]:
        values = pd.to_numeric(work[col], errors="coerce")
        daily[str(col)] = values.groupby(work["date"]).sum(min_count=1)
    daily = daily.reset_index().sort_values("date")
    return name, daily


def choose_targets(daily: pd.DataFrame) -> list[str]:
    blocked_words = ["id", "serial", "status", "way", "upload", "wallet", "phone", "qr", "card", "reduction", "编号", "状态", "方式", "折扣"]
    candidates = []
    for c in daily.columns:
        low = str(c).lower()
        if c != "date" and not any(word in low for word in blocked_words):
            candidates.append(c)
    preferred_words = [
        "record_count",
        "人数",
        "数量",
        "销量",
        "包裹量",
        "consume_money",
        "销售",
        "金额",
        "revenue",
        "calories",
        "carbohydrates",
        "protein",
        "fat",
        "fiber",
    ]
    ordered = []
    for word in preferred_words:
        for col in candidates:
            if col not in ordered and word.lower() in str(col).lower():
                ordered.append(col)
    for col in candidates:
        if col not in ordered:
            ordered.append(col)
    valid = []
    for col in ordered:
        values = pd.to_numeric(daily[col], errors="coerce")
        if values.notna().sum() >= 30 and values.std(skipna=True) > 0:
            valid.append(col)
    return valid[:8]


def make_features(daily: pd.DataFrame, targets: list[str]) -> pd.DataFrame:
    df = daily.copy().sort_values("date").reset_index(drop=True)
    df["date"] = pd.to_datetime(df["date"])
    df["ord"] = (df["date"] - df["date"].min()).dt.days
    df["dow"] = df["date"].dt.dayofweek
    df["month"] = df["date"].dt.month
    df["day"] = df["date"].dt.day
    df["sin_doy"] = np.sin(2 * np.pi * df["date"].dt.dayofyear / 365.25)
    df["cos_doy"] = np.cos(2 * np.pi * df["date"].dt.dayofyear / 365.25)
    for target in targets:
        values = pd.to_numeric(df[target], errors="coerce")
        df[target] = values
        for lag in [1, 5, 20]:
            df[f"{target}_lag{lag}"] = values.shift(lag)
        for win in [5, 20]:
            df[f"{target}_roll{win}"] = values.shift(1).rolling(win, min_periods=3).mean()
    return df.dropna().reset_index(drop=True)


def metrics(y_true: pd.Series, pred: np.ndarray) -> dict:
    pred = np.maximum(np.asarray(pred, dtype=float), 0)
    y = np.asarray(y_true, dtype=float)
    rmse = math.sqrt(mean_squared_error(y, pred))
    mae = mean_absolute_error(y, pred)
    denom = np.maximum(np.abs(y), 1e-6)
    mape = float(np.mean(np.abs((y - pred) / denom)) * 100)
    return {"RMSE": rmse, "MAE": mae, "MAPE(%)": mape, "R2": r2_score(y, pred)}


def model_candidates() -> dict:
    return {
        "Ridge": make_pipeline(StandardScaler(), Ridge(alpha=3.0)),
        "RandomForest": RandomForestRegressor(n_estimators=240, max_depth=10, min_samples_leaf=3, random_state=20260520, n_jobs=-1),
        "HistGradientBoosting": HistGradientBoostingRegressor(max_iter=260, learning_rate=0.05, l2_regularization=0.05, random_state=20260520),
    }


def future_business_dates(last_date: pd.Timestamp, horizon: int = 20) -> pd.DatetimeIndex:
    start = pd.Timestamp(last_date) + pd.Timedelta(days=1)
    return pd.bdate_range(start, periods=horizon)


def run_prediction_model(daily_source: str, daily: pd.DataFrame, manifest: dict) -> None:
    targets = choose_targets(daily)
    if not targets:
        manifest["notes"].append("没有找到可预测的数值目标。")
        return
    data = make_features(daily, targets)
    if len(data) < 50:
        manifest["notes"].append("可用于滚动验证的数据少于 50 行，跳过预测建模。")
        return
    raw_daily_cols = [c for c in daily.columns if c != "date"]
    feature_cols = [c for c in data.columns if c not in ["date", *raw_daily_cols]]
    valid_n = min(45, max(12, int(len(data) * 0.18)))
    train = data.iloc[:-valid_n]
    valid = data.iloc[-valid_n:]
    metric_rows = []
    best_models = {}
    best_names = {}
    for target in targets:
        best_rmse = float("inf")
        best_name = ""
        for name, model in model_candidates().items():
            model.fit(train[feature_cols], train[target])
            pred = model.predict(valid[feature_cols])
            row = {"target": target, "model": name, **metrics(valid[target], pred)}
            metric_rows.append(row)
            if row["RMSE"] < best_rmse:
                best_rmse = row["RMSE"]
                best_name = name
        final_model = model_candidates()[best_name]
        final_model.fit(data[feature_cols], data[target])
        best_models[target] = final_model
        best_names[target] = best_name
    metric_df = pd.DataFrame(metric_rows).sort_values(["target", "RMSE"])
    metric_path = SPECIAL_DIR / "prediction_validation_metrics.csv"
    metric_df.to_csv(metric_path, index=False, encoding="utf-8-sig")
    manifest["tables"].append(metric_path.relative_to(ROOT).as_posix())

    history = daily[["date", *targets]].copy().sort_values("date").reset_index(drop=True)
    forecasts = []
    for date in future_business_dates(history["date"].max(), 20):
        feat = add_future_feature(history, date, targets, daily["date"].min())
        row = {"date": date}
        for target in targets:
            pred = float(np.maximum(best_models[target].predict(feat[feature_cols])[0], 0))
            row[target] = pred
        forecasts.append(row)
        history = pd.concat([history, pd.DataFrame([row])], ignore_index=True)
    forecast_df = pd.DataFrame(forecasts)
    forecast_path = SPECIAL_DIR / "prediction_future_workdays.csv"
    forecast_df.to_csv(forecast_path, index=False, encoding="utf-8-sig")
    manifest["tables"].append(forecast_path.relative_to(ROOT).as_posix())

    daily_path = SPECIAL_DIR / "prediction_daily_dataset.csv"
    daily.to_csv(daily_path, index=False, encoding="utf-8-sig")
    manifest["tables"].append(daily_path.relative_to(ROOT).as_posix())

    first_target = targets[0]
    fig, ax = plt.subplots(figsize=(10, 4.5))
    ax.plot(daily["date"], daily[first_target], color="#657484", linewidth=1.0, alpha=0.7, label="历史")
    ax.plot(forecast_df["date"], forecast_df[first_target], color="#2358a6", marker="o", linewidth=1.8, label="预测")
    ax.set_xlabel("日期")
    ax.set_ylabel(str(first_target))
    ax.legend()
    fig.tight_layout()
    fig_path = FIG_DIR / "specialized_prediction_forecast.png"
    fig.savefig(fig_path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    manifest["figures"].append(fig_path.relative_to(ROOT).as_posix())
    manifest["specialized_models"].append(
        {
            "type": "prediction",
            "source": daily_source,
            "targets": targets,
            "best_models": best_names,
            "validation_rows": int(valid_n),
            "forecast_horizon": 20,
        }
    )


def add_future_feature(history: pd.DataFrame, date: pd.Timestamp, targets: list[str], start_date: pd.Timestamp) -> pd.DataFrame:
    row = {
        "date": date,
        "ord": (pd.Timestamp(date) - pd.Timestamp(start_date)).days,
        "dow": date.dayofweek,
        "month": date.month,
        "day": date.day,
        "sin_doy": math.sin(2 * math.pi * date.dayofyear / 365.25),
        "cos_doy": math.cos(2 * math.pi * date.dayofyear / 365.25),
    }
    for target in targets:
        values = pd.to_numeric(history[target], errors="coerce").ffill().fillna(0)
        for lag in [1, 5, 20]:
            row[f"{target}_lag{lag}"] = float(values.iloc[-lag]) if len(values) >= lag else float(values.iloc[-1])
        for win in [5, 20]:
            row[f"{target}_roll{win}"] = float(values.tail(win).mean())
    return pd.DataFrame([row])


def classify_dish(name: str) -> str:
    text = str(name)
    if any(k in text for k in ["米饭", "饭", "面", "粉", "包", "馒头", "饼"]):
        return "主食"
    if any(k in text for k in ["鱼", "虾", "蟹", "带鱼"]):
        return "水产"
    if any(k in text for k in ["鸡", "鸭", "肉", "排", "牛", "猪", "肠"]):
        return "肉禽"
    if "蛋" in text:
        return "蛋类"
    if any(k in text for k in ["豆腐", "豆", "腐", "千张"]):
        return "豆制品"
    if any(k in text for k in ["菜", "笋", "瓜", "菇", "木耳", "花菜", "萝卜", "茄子", "土豆", "山药", "芋头"]):
        return "蔬菜"
    return "其他"


def run_dish_analysis(tables: list[tuple[str, pd.DataFrame]], manifest: dict) -> None:
    dish_tables = []
    for name, df in merge_compatible_tables(tables):
        cols = {str(c).lower(): c for c in df.columns}
        if "dish_name" in cols or "菜品" in "".join(map(str, df.columns)):
            if any("weight" in str(c).lower() or "重量" in str(c) for c in df.columns):
                dish_tables.append((name, normalize_columns(df)))
    if not dish_tables:
        return
    name, details = max(dish_tables, key=lambda item: len(item[1]))
    dish_col = pick_col(details, ["dish_name", "菜品", "菜名"])
    weight_col = pick_col(details, ["weight", "重量"])
    price_col = pick_col(details, ["total_price", "price", "金额", "价格"])
    order_col = pick_col(details, ["indent_id", "order", "订单"])
    if not dish_col or not weight_col or not price_col:
        return
    details[weight_col] = pd.to_numeric(details[weight_col], errors="coerce")
    details[price_col] = pd.to_numeric(details[price_col], errors="coerce")
    details = details[details[weight_col].gt(0) & details[price_col].ge(0)]
    group = {
        "total_weight": (weight_col, "sum"),
        "total_sales": (price_col, "sum"),
        "median_weight": (weight_col, "median"),
        "median_price": (price_col, "median"),
        "line_count": (dish_col, "count"),
    }
    if order_col:
        group["order_count"] = (order_col, "nunique")
    profile = details.groupby(dish_col).agg(**group).reset_index().rename(columns={dish_col: "dish_name"})
    if "order_count" not in profile.columns:
        profile["order_count"] = profile["line_count"]
    profile["category"] = profile["dish_name"].map(classify_dish)
    profile["price_per_g"] = profile["total_sales"] / profile["total_weight"].replace(0, np.nan)
    profile = profile.sort_values("total_weight", ascending=False)
    path = SPECIAL_DIR / "dish_profile.csv"
    profile.to_csv(path, index=False, encoding="utf-8-sig")
    manifest["tables"].append(path.relative_to(ROOT).as_posix())

    top = profile.head(20)
    fig, ax = plt.subplots(figsize=(10, 5.6))
    ax.barh(top["dish_name"][::-1], top["total_weight"][::-1], color="#237a57")
    ax.set_xlabel("累计重量")
    fig.tight_layout()
    fig_path = FIG_DIR / "specialized_top_dishes.png"
    fig.savefig(fig_path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    manifest["figures"].append(fig_path.relative_to(ROOT).as_posix())

    package = build_package_plan(profile)
    if not package.empty:
        package_path = SPECIAL_DIR / "heuristic_package_plan.csv"
        package.to_csv(package_path, index=False, encoding="utf-8-sig")
        manifest["tables"].append(package_path.relative_to(ROOT).as_posix())
    meal_plan = build_meal_preparation_plan(profile)
    if not meal_plan.empty:
        meal_path = SPECIAL_DIR / "meal_preparation_plan.csv"
        meal_plan.to_csv(meal_path, index=False, encoding="utf-8-sig")
        manifest["tables"].append(meal_path.relative_to(ROOT).as_posix())
        meal_summary = meal_plan.groupby(["date", "meal"], as_index=False).agg(total_kg=("planned_kg", "sum"), dishes=("dish_name", "nunique"))
        fig, ax = plt.subplots(figsize=(9, 4.5))
        meals = list(meal_summary["meal"].drop_duplicates())
        x = np.arange(meal_summary["date"].nunique())
        width = 0.35
        dates = list(meal_summary["date"].drop_duplicates())
        for idx, meal in enumerate(meals):
            values = []
            for date in dates:
                row = meal_summary[(meal_summary["date"].eq(date)) & (meal_summary["meal"].eq(meal))]
                values.append(float(row["total_kg"].iloc[0]) if not row.empty else 0.0)
            ax.bar(x + (idx - (len(meals) - 1) / 2) * width, values, width=width, label=meal)
        ax.set_xticks(x)
        ax.set_xticklabels(dates, rotation=25, ha="right")
        ax.set_ylabel("备菜量/kg")
        ax.legend()
        fig.tight_layout()
        fig_path = FIG_DIR / "specialized_meal_plan_total.png"
        fig.savefig(fig_path, dpi=300, bbox_inches="tight")
        plt.close(fig)
        manifest["figures"].append(fig_path.relative_to(ROOT).as_posix())
        manifest["specialized_models"].append({"type": "meal_preparation", "source": "prediction_future_workdays.csv + dish_profile.csv", "rows": int(len(meal_plan))})
    manifest["specialized_models"].append({"type": "dish_optimization", "source": name, "dish_count": int(len(profile))})


def pick_col(df: pd.DataFrame, keywords: list[str]) -> str | None:
    for col in df.columns:
        low = str(col).lower()
        if any(key.lower() in low for key in keywords):
            return str(col)
    return None


def build_package_plan(profile: pd.DataFrame) -> pd.DataFrame:
    candidates = profile[(profile["median_price"] > 0) & (profile["median_weight"] > 0) & (profile["order_count"] >= 5)].copy()
    if candidates.empty:
        return pd.DataFrame()
    packages = []
    def pool_frame(frame: pd.DataFrame, n_pop: int, n_price: int, cap: int) -> pd.DataFrame:
        return pd.concat(
            [frame.sort_values("order_count", ascending=False).head(n_pop), frame.sort_values("median_price", ascending=False).head(n_price)]
        ).drop_duplicates("dish_name").head(cap)

    pools = {
        "主食": pool_frame(candidates[candidates["category"].eq("主食")], 4, 2, 5),
        "蛋白": pool_frame(candidates[candidates["category"].isin(["肉禽", "水产", "蛋类", "豆制品"])], 7, 7, 10),
        "蔬菜": pool_frame(candidates[candidates["category"].eq("蔬菜")], 6, 5, 8),
    }
    forms = {
        10: [(1, 1, 2), (1, 2, 1)],
        15: [(1, 2, 2), (1, 3, 1)],
        20: [(1, 3, 2), (1, 4, 1)],
    }
    for level in [10, 15, 20]:
        best = None
        for staple_n, protein_n, veg_n in forms[level]:
            for staples in itertools.combinations(pools["主食"].to_dict("records"), staple_n):
                for proteins in itertools.combinations(pools["蛋白"].to_dict("records"), protein_n):
                    for vegs in itertools.combinations(pools["蔬菜"].to_dict("records"), veg_n):
                        items = list(staples) + list(proteins) + list(vegs)
                        if len({x.get("dish_name") for x in items}) != len(items):
                            continue
                        price = sum(float(x.get("median_price", 0)) for x in items)
                        if price <= 0:
                            continue
                        price_penalty = abs(price - level) / level
                        pop_bonus = np.mean([math.log1p(float(x.get("order_count", 0))) for x in items]) / 100
                        category_bonus = len({x.get("category") for x in items}) / 100
                        score = price_penalty - pop_bonus - category_bonus
                        if best is None or score < best["score"]:
                            best = {"items": items, "price": price, "score": score}
        if best:
            for item in best["items"]:
                packages.append(
                    {
                        "price_level": f"{level}元",
                        "dish_name": item["dish_name"],
                        "category": item["category"],
                        "portion_g": round(float(item["median_weight"]), 1),
                        "portion_price": round(float(item["median_price"]), 2),
                    }
                )
            packages.append(
                {
                    "price_level": f"{level}元合计",
                    "dish_name": "合计",
                    "category": "-",
                    "portion_g": round(sum(float(x["median_weight"]) for x in best["items"]), 1),
                    "portion_price": round(best["price"], 2),
                }
            )
    return pd.DataFrame(packages)


def build_meal_preparation_plan(profile: pd.DataFrame) -> pd.DataFrame:
    forecast_path = SPECIAL_DIR / "prediction_future_workdays.csv"
    if not forecast_path.exists() or profile.empty:
        return pd.DataFrame()
    forecast = pd.read_csv(forecast_path)
    if "date" not in forecast.columns:
        return pd.DataFrame()
    forecast["date"] = pd.to_datetime(forecast["date"], errors="coerce")
    forecast = forecast[forecast["date"].notna()]
    target_dates = pd.to_datetime(["2025-05-06", "2025-05-07", "2025-05-08", "2025-05-09", "2025-05-12"])
    forecast = forecast[forecast["date"].isin(target_dates)]
    if forecast.empty:
        forecast = pd.read_csv(forecast_path).head(5)
        forecast["date"] = pd.to_datetime(forecast["date"], errors="coerce")
    demand_col = "record_count" if "record_count" in forecast.columns else None
    if not demand_col:
        numeric_cols = forecast.select_dtypes(include=[np.number]).columns.tolist()
        if not numeric_cols:
            return pd.DataFrame()
        demand_col = numeric_cols[0]
    candidates = profile[(profile["median_weight"] > 0) & (profile["total_weight"] > 0)].copy()
    if candidates.empty:
        return pd.DataFrame()
    selected = select_menu(candidates, n=14)
    dinner_selected = select_menu(candidates, n=8)
    basket_grams = float(np.clip(profile["median_weight"].median() * 7, 360, 520))
    rows = []
    for _, day in forecast.iterrows():
        diners = max(float(day[demand_col]), 1.0)
        for meal, share, safety, menu in [("午餐", 0.92, 1.08, selected), ("晚餐", 0.08, 1.12, dinner_selected)]:
            total_grams = diners * share * basket_grams * safety
            weights = menu["total_weight"].astype(float).to_numpy()
            weights = weights / weights.sum()
            for (_, dish), dish_share in zip(menu.iterrows(), weights):
                grams = total_grams * float(dish_share)
                rows.append(
                    {
                        "date": pd.Timestamp(day["date"]).date().isoformat(),
                        "meal": meal,
                        "dish_name": dish["dish_name"],
                        "category": dish["category"],
                        "planned_kg": round(grams / 1000, 2),
                        "expected_servings": int(round(grams / max(float(dish["median_weight"]), 1.0))),
                        "estimated_sales": round(grams * float(dish.get("price_per_g", 0.0)), 2),
                    }
                )
    return pd.DataFrame(rows)


def select_menu(profile: pd.DataFrame, n: int) -> pd.DataFrame:
    selected = []
    for category in ["主食", "肉禽", "水产", "蛋类", "豆制品", "蔬菜"]:
        subset = profile[profile["category"].eq(category)].sort_values("total_weight", ascending=False)
        if not subset.empty:
            selected.append(str(subset.iloc[0]["dish_name"]))
    remaining = profile[~profile["dish_name"].isin(selected)].sort_values("total_weight", ascending=False)
    selected.extend(remaining.head(max(0, n - len(selected)))["dish_name"].tolist())
    return profile[profile["dish_name"].isin(selected)].drop_duplicates("dish_name").head(n)


def write_summary(manifest: dict) -> None:
    lines = [
        "# 专项建模运行摘要",
        "",
        f"- 推荐赛题：{PROBLEM_ID}",
        f"- 赛题标题：{PROBLEM_TITLE}",
        f"- 专项模型数量：{len(manifest['specialized_models'])}",
        f"- 输出表格数量：{len(manifest['tables'])}",
        f"- 输出图片数量：{len(manifest['figures'])}",
        "",
        "## 模型清单",
    ]
    for model in manifest["specialized_models"]:
        lines.append(f"- {model}")
    if manifest["notes"]:
        lines.extend(["", "## 注意事项"])
        lines.extend([f"- {note}" for note in manifest["notes"]])
    (RESULTS_DIR / "specialized_summary.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    ensure_dirs()
    tables = load_tables()
    manifest = {
        "problem_id": PROBLEM_ID,
        "problem_title": PROBLEM_TITLE,
        "specialized_models": [],
        "tables": [],
        "figures": [],
        "notes": [],
    }
    daily_bundle = build_daily_dataset(tables)
    if daily_bundle:
        source, daily = daily_bundle
        run_prediction_model(source, daily, manifest)
    else:
        manifest["notes"].append("未找到适合时间序列预测的日期字段与数值字段。")
    run_dish_analysis(tables, manifest)
    if not manifest["specialized_models"]:
        manifest["notes"].append("当前题型尚未匹配到专项模型，已保留基线数据画像作为主要结果。")
    write_summary(manifest)
    manifest["summary_markdown"] = "results/specialized_summary.md"
    (RESULTS_DIR / "specialized_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"ok": True, "models": len(manifest["specialized_models"]), "tables": len(manifest["tables"]), "figures": len(manifest["figures"])}, ensure_ascii=False))


if __name__ == "__main__":
    main()
'''
    return template.replace("__CONFIG_JSON__", payload_json)
