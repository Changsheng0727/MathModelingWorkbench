from __future__ import annotations

from pathlib import Path
from typing import Any

from app.services.executor import run_python_script
from app.services.store import save_json


def generate_modeling_script(root: Path, analysis: dict[str, Any]) -> dict[str, str]:
    code_dir = root / "code"
    results_dir = root / "results"
    code_dir.mkdir(parents=True, exist_ok=True)
    results_dir.mkdir(parents=True, exist_ok=True)
    script_path = code_dir / "run_baseline_analysis.py"
    script_path.write_text(render_script(analysis), encoding="utf-8")
    return {"modeling_script": "code/run_baseline_analysis.py"}


def run_modeling_script(root: Path, timeout: int = 240) -> dict[str, Any]:
    script_path = root / "code" / "run_baseline_analysis.py"
    if not script_path.exists():
        raise FileNotFoundError("请先生成建模脚本")

    run_result = run_python_script(root, "code/run_baseline_analysis.py", "artifacts/modeling_run.log", timeout)

    manifest_path = root / "results" / "baseline_manifest.json"
    manifest = {}
    if manifest_path.exists():
        try:
            import json

            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        except Exception:
            manifest = {}
    payload = {
        "success": run_result["success"],
        "returncode": run_result["returncode"],
        "executor": run_result["executor"],
        "log": run_result["log"],
        "manifest": "results/baseline_manifest.json" if manifest_path.exists() else "",
        "outputs": manifest,
    }
    save_json(root / "artifacts" / "modeling_status.json", payload)
    return payload


def render_script(analysis: dict[str, Any]) -> str:
    recommended = analysis.get("recommended_problem", {})
    problem_id = recommended.get("id", "Unknown")
    title = recommended.get("title", "Unknown problem")
    return f'''from __future__ import annotations

import json
import math
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook

plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "Arial Unicode MS", "DejaVu Sans"]
plt.rcParams["axes.unicode_minus"] = False


ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "raw"
RESULTS_DIR = ROOT / "results"
TABLE_DIR = RESULTS_DIR / "tables"
FIG_DIR = RESULTS_DIR / "figures"
PROBLEM_ID = {problem_id!r}
PROBLEM_TITLE = {title!r}


def ensure_dirs() -> None:
    TABLE_DIR.mkdir(parents=True, exist_ok=True)
    FIG_DIR.mkdir(parents=True, exist_ok=True)


def read_csv(path: Path) -> pd.DataFrame:
    for encoding in ["utf-8-sig", "gbk", "utf-8"]:
        try:
            return pd.read_csv(path, encoding=encoding)
        except Exception:
            continue
    raise ValueError(f"cannot read csv: {{path}}")


def iter_tables() -> list[tuple[str, pd.DataFrame]]:
    tables = []
    all_files = [path for path in sorted(RAW_DIR.rglob("*")) if path.is_file() and path.suffix.lower() in [".csv", ".xlsx", ".xls"]]
    selected = []
    for path in all_files:
        rel = path.relative_to(RAW_DIR).as_posix()
        if PROBLEM_ID != "Unknown" and (
            f"赛题{{PROBLEM_ID}}" in rel
            or f"Problem {{PROBLEM_ID}}" in rel
            or f"题{{PROBLEM_ID}}" in rel
            or f"{{PROBLEM_ID}}-" in rel
            or f"{{PROBLEM_ID}}_" in rel
        ):
            selected.append(path)
    files = selected or all_files
    for path in files:
        if not path.is_file():
            continue
        rel = path.relative_to(RAW_DIR).as_posix()
        suffix = path.suffix.lower()
        try:
            if suffix == ".csv":
                df = read_csv(path)
                tables.append((rel, df))
            elif suffix in [".xlsx", ".xls"]:
                workbook = load_workbook(path, read_only=True, data_only=True)
                for sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    if ws.max_row <= 1 or ws.max_column <= 0:
                        continue
                    # Avoid loading very large sheets blindly beyond a practical MVP cap.
                    nrows = min(ws.max_row, 200000)
                    df = pd.read_excel(path, sheet_name=sheet_name, nrows=nrows)
                    tables.append((f"{{rel}}::{{sheet_name}}", df))
                workbook.close()
        except Exception as exc:
            print(f"[WARN] failed to read {{rel}}: {{type(exc).__name__}} {{exc}}")
    return tables


def missing_profile(name: str, df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for col in df.columns:
        series = df[col]
        rows.append(
            {{
                "table": name,
                "column": str(col),
                "dtype": str(series.dtype),
                "missing_count": int(series.isna().sum()),
                "missing_rate": float(series.isna().mean()),
                "unique_count": int(series.nunique(dropna=True)),
            }}
        )
    return pd.DataFrame(rows)


def numeric_profile(name: str, df: pd.DataFrame) -> pd.DataFrame:
    numeric = df.select_dtypes(include=[np.number])
    if numeric.empty:
        return pd.DataFrame()
    desc = numeric.describe(percentiles=[0.25, 0.5, 0.75]).T.reset_index()
    desc.insert(0, "table", name)
    desc = desc.rename(columns={{"index": "column"}})
    return desc


def find_datetime_column(df: pd.DataFrame) -> str | None:
    candidates = []
    for col in df.columns:
        text = str(col).lower()
        if any(key in text for key in ["date", "time", "日期", "时间"]):
            candidates.append(col)
    candidates.extend([col for col in df.columns if col not in candidates])
    sample = df.head(2000)
    for col in candidates:
        parsed = pd.to_datetime(sample[col], errors="coerce")
        if parsed.notna().mean() >= 0.6:
            return str(col)
    return None


def plot_numeric_histograms(name: str, df: pd.DataFrame, safe_name: str) -> list[str]:
    numeric = df.select_dtypes(include=[np.number])
    outputs = []
    if numeric.empty:
        return outputs
    cols = list(numeric.columns[:6])
    rows = math.ceil(len(cols) / 2)
    fig, axes = plt.subplots(rows, 2, figsize=(10, max(3.2, rows * 3.0)))
    axes = np.array(axes).reshape(-1)
    for ax, col in zip(axes, cols):
        values = pd.to_numeric(numeric[col], errors="coerce").dropna()
        values = values[np.isfinite(values)]
        if len(values) > 5000:
            values = values.sample(5000, random_state=20260520)
        ax.hist(values, bins=32, color="#2358a6", alpha=0.82)
        ax.set_title(str(col))
    for ax in axes[len(cols):]:
        ax.axis("off")
    fig.suptitle(f"数值变量分布：{{name}}")
    fig.tight_layout()
    path = FIG_DIR / f"{{safe_name}}_numeric_hist.png"
    fig.savefig(path, dpi=180)
    plt.close(fig)
    outputs.append(path.relative_to(ROOT).as_posix())
    return outputs


def plot_correlation(name: str, df: pd.DataFrame, safe_name: str) -> list[str]:
    numeric = df.select_dtypes(include=[np.number])
    outputs = []
    if numeric.shape[1] < 2:
        return outputs
    corr = numeric.iloc[:, :12].corr(numeric_only=True)
    fig, ax = plt.subplots(figsize=(8, 6.5))
    im = ax.imshow(corr, cmap="RdBu_r", vmin=-1, vmax=1)
    ax.set_xticks(range(len(corr.columns)))
    ax.set_yticks(range(len(corr.index)))
    ax.set_xticklabels([str(c) for c in corr.columns], rotation=45, ha="right", fontsize=8)
    ax.set_yticklabels([str(c) for c in corr.index], fontsize=8)
    fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    ax.set_title(f"数值变量相关性：{{name}}")
    fig.tight_layout()
    path = FIG_DIR / f"{{safe_name}}_correlation.png"
    fig.savefig(path, dpi=180)
    plt.close(fig)
    corr.to_csv(TABLE_DIR / f"{{safe_name}}_correlation.csv", encoding="utf-8-sig")
    outputs.append(path.relative_to(ROOT).as_posix())
    return outputs


def plot_time_series(name: str, df: pd.DataFrame, safe_name: str) -> list[str]:
    date_col = find_datetime_column(df)
    numeric = df.select_dtypes(include=[np.number])
    outputs = []
    if not date_col or numeric.empty:
        return outputs
    value_candidates = [col for col in numeric.columns if str(col) != str(date_col)]
    if not value_candidates:
        return outputs
    work = df.copy()
    work[date_col] = pd.to_datetime(work[date_col], errors="coerce")
    work = work[work[date_col].notna()]
    if work.empty:
        return outputs
    value_col = value_candidates[0]
    work[value_col] = pd.to_numeric(work[value_col], errors="coerce")
    work = work[work[value_col].notna()]
    if work.empty:
        return outputs
    daily = work.groupby(work[date_col].dt.date)[value_col].sum().reset_index()
    daily.columns = ["date", str(value_col)]
    daily["date"] = pd.to_datetime(daily["date"])
    daily.to_csv(TABLE_DIR / f"{{safe_name}}_daily_series.csv", index=False, encoding="utf-8-sig")
    fig, ax = plt.subplots(figsize=(10, 4.5))
    ax.plot(daily["date"], daily[str(value_col)], color="#237a57", linewidth=1.4)
    ax.set_title(f"按日聚合趋势：{{name}} / {{value_col}}")
    ax.set_xlabel("日期")
    ax.set_ylabel(str(value_col))
    fig.tight_layout()
    path = FIG_DIR / f"{{safe_name}}_daily_trend.png"
    fig.savefig(path, dpi=180)
    plt.close(fig)
    outputs.append(path.relative_to(ROOT).as_posix())
    return outputs


def safe_stem(name: str) -> str:
    keep = []
    for ch in name:
        if ch.isalnum() or ch in ["-", "_"]:
            keep.append(ch)
        else:
            keep.append("_")
    stem = "".join(keep).strip("_")
    return stem[:80] or "table"


def write_summary(summary: dict) -> None:
    lines = [
        "# 基线建模运行摘要",
        "",
        f"- 推荐赛题：{{PROBLEM_ID}}",
        f"- 赛题标题：{{PROBLEM_TITLE}}",
        f"- 解析数据表数量：{{summary['table_count']}}",
        f"- 生成图表数量：{{len(summary['figures'])}}",
        f"- 生成结果表数量：{{len(summary['tables'])}}",
        "",
        "## 数据表概览",
    ]
    for item in summary["tables_overview"]:
        lines.append(
            f"- {{item['name']}}：{{item['rows']}} 行，{{item['cols']}} 列，"
            f"数值列 {{item['numeric_cols']}} 个，缺失率均值 {{item['mean_missing_rate']:.2%}}"
        )
    lines.extend(["", "## 后续建议"])
    lines.append("- 若题目包含预测任务，应在此基础上构造时间特征、训练基线模型并进行滚动验证。")
    lines.append("- 若题目包含优化任务，应将清洗后的统计量转换为目标函数、约束和可行域。")
    lines.append("- 论文中的图表应从本目录结果引用，并在正文中补充自然判读段落。")
    (RESULTS_DIR / "baseline_summary.md").write_text("\\n".join(lines), encoding="utf-8")


def main() -> None:
    ensure_dirs()
    tables = iter_tables()
    all_missing = []
    all_numeric = []
    summary = {{
        "problem_id": PROBLEM_ID,
        "problem_title": PROBLEM_TITLE,
        "table_count": len(tables),
        "tables_overview": [],
        "figures": [],
        "tables": [],
    }}
    for idx, (name, df) in enumerate(tables, 1):
        safe_name = f"table_{{idx:02d}}_{{safe_stem(name)}}"
        missing = missing_profile(name, df)
        numeric = numeric_profile(name, df)
        all_missing.append(missing)
        if not numeric.empty:
            all_numeric.append(numeric)
        missing.to_csv(TABLE_DIR / f"{{safe_name}}_missing_profile.csv", index=False, encoding="utf-8-sig")
        summary["tables"].append(f"results/tables/{{safe_name}}_missing_profile.csv")
        if not numeric.empty:
            numeric.to_csv(TABLE_DIR / f"{{safe_name}}_numeric_profile.csv", index=False, encoding="utf-8-sig")
            summary["tables"].append(f"results/tables/{{safe_name}}_numeric_profile.csv")
        summary["figures"].extend(plot_numeric_histograms(name, df, safe_name))
        summary["figures"].extend(plot_correlation(name, df, safe_name))
        summary["figures"].extend(plot_time_series(name, df, safe_name))
        summary["tables_overview"].append(
            {{
                "name": name,
                "rows": int(len(df)),
                "cols": int(df.shape[1]),
                "numeric_cols": int(df.select_dtypes(include=[np.number]).shape[1]),
                "mean_missing_rate": float(df.isna().mean().mean()) if df.shape[1] else 0.0,
            }}
        )
    if all_missing:
        merged_missing = pd.concat(all_missing, ignore_index=True)
        merged_missing.to_csv(TABLE_DIR / "all_missing_profile.csv", index=False, encoding="utf-8-sig")
        summary["tables"].append("results/tables/all_missing_profile.csv")
    if all_numeric:
        merged_numeric = pd.concat(all_numeric, ignore_index=True)
        merged_numeric.to_csv(TABLE_DIR / "all_numeric_profile.csv", index=False, encoding="utf-8-sig")
        summary["tables"].append("results/tables/all_numeric_profile.csv")
    write_summary(summary)
    summary["summary_markdown"] = "results/baseline_summary.md"
    (RESULTS_DIR / "baseline_manifest.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps({{"ok": True, "tables": len(tables), "figures": len(summary["figures"])}}, ensure_ascii=False))


if __name__ == "__main__":
    main()
'''
